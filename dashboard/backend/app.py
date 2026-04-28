from fastapi import FastAPI, Depends, HTTPException, Header, Request
import logging
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from typing import Optional, List
import os
import threading
from pathlib import Path
import numpy as np
import pandas as pd
import joblib
from typing import Any
try:
    from openpyxl import Workbook, load_workbook
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

try:
    import tensorflow as tf
    from tensorflow.keras.models import load_model as keras_load_model
    TF_AVAILABLE = True
except Exception:
    tf = None
    keras_load_model = None
    TF_AVAILABLE = False

from . import models
from .database import SessionLocal, init_db, engine
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from pathlib import Path
from fastapi.responses import StreamingResponse
import asyncio
import json
from datetime import datetime, timezone
from pydantic import BaseModel

# initialize DB (models are imported above so metadata is available)
init_db()

app = FastAPI(title="Soil Sensor Dashboard API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

API_KEY = os.getenv("DASHBOARD_API_KEY", "dev-token")
LEGACY_API_KEY = "dev-token"

# Prediction artifacts (loaded lazily on first prediction request)
PREDICTION_ROOT = Path(__file__).resolve().parents[2]
PREDICTION_MODEL_DIR = PREDICTION_ROOT / "model"
prediction_scaler = None
prediction_label_encoder = None
prediction_keras_model = None
prediction_sklearn_model = None
prediction_loaded = False
PREDICTION_FEATURES = ["Nitrogen", "Phosphorus", "Potassium", "pH", "Humidity", "Temperature"]


def _existing_path(candidates: List[Path]) -> Optional[Path]:
    for p in candidates:
        if p.exists():
            return p
    return None


def _load_prediction_artifacts() -> None:
    global prediction_scaler, prediction_label_encoder
    global prediction_keras_model, prediction_sklearn_model, prediction_loaded

    if prediction_loaded:
        return

    model_dir = PREDICTION_MODEL_DIR / "chili_eggplant_model"
    scaler_path = model_dir / "scaler.save"
    encoder_path = model_dir / "label_encoder.save"

    if not scaler_path.exists() or not encoder_path.exists():
        raise RuntimeError("Missing Chili/Eggplant scaler or label encoder artifacts")

    prediction_scaler = joblib.load(scaler_path)
    prediction_label_encoder = joblib.load(encoder_path)

    if TF_AVAILABLE and keras_load_model is not None:
        for model_path in [
            model_dir / "chili_eggplant_model.keras",
            model_dir / "chili_eggplant_model.h5",
            model_dir / "best_model.keras",
        ]:
            if not model_path.exists():
                continue
            try:
                prediction_keras_model = keras_load_model(str(model_path))
                logging.info("Loaded prediction model from %s", model_path)
                prediction_loaded = True
                return
            except Exception as e:
                logging.warning("Could not load Keras model %s: %s", model_path, e)

    raise RuntimeError("No Chili/Eggplant Keras model artifact could be loaded")


def _predict_crop(values: List[float]) -> dict:
    _load_prediction_artifacts()

    if prediction_scaler is None or prediction_label_encoder is None:
        raise RuntimeError("Prediction preprocessors are not loaded")

    feature_names = getattr(prediction_scaler, "feature_names_in_", None)
    if feature_names is not None and len(feature_names) == len(values):
        feature_frame = pd.DataFrame([values], columns=list(feature_names))
        scaled = prediction_scaler.transform(feature_frame)
    else:
        feature_array = np.array([values], dtype=float)
        scaled = prediction_scaler.transform(feature_array)
    class_names = [str(c) for c in prediction_label_encoder.classes_]
    probs = None

    if prediction_keras_model is not None:
        raw = prediction_keras_model.predict(scaled, verbose=0)
        probs = np.array(raw[0], dtype=float)
    elif prediction_sklearn_model is not None:
        if hasattr(prediction_sklearn_model, "predict_proba"):
            probs = np.array(prediction_sklearn_model.predict_proba(scaled)[0], dtype=float)
            if hasattr(prediction_sklearn_model, "classes_"):
                sk_classes = [str(c) for c in prediction_sklearn_model.classes_]
                if len(sk_classes) == len(probs):
                    class_names = sk_classes
        else:
            pred = prediction_sklearn_model.predict(scaled)
            pred_label = str(pred[0])
            return {
                "predicted_crop": pred_label,
                "confidence": 1.0,
                "top_predictions": [{"crop": pred_label, "probability": 1.0}],
            }

    if probs is None or probs.size == 0:
        raise RuntimeError("Prediction model returned no probabilities")

    top_idx = int(np.argmax(probs))
    sorted_idx = np.argsort(probs)[::-1][:3]
    top_predictions = [
        {
            "crop": class_names[int(i)] if int(i) < len(class_names) else str(int(i)),
            "probability": float(probs[int(i)]),
        }
        for i in sorted_idx
    ]
    predicted_crop = class_names[top_idx] if top_idx < len(class_names) else str(top_idx)
    confidence = float(probs[top_idx])

    return {
        "predicted_crop": predicted_crop,
        "confidence": confidence,
        "top_predictions": top_predictions,
    }


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def require_api_key(x_api_key: Optional[str] = Header(None)):
    if x_api_key is None:
        raise HTTPException(status_code=401, detail="Missing API key")
    if x_api_key not in {API_KEY, LEGACY_API_KEY}:
        raise HTTPException(status_code=403, detail="Invalid API key")


# Simple in-memory broadcaster for Server-Sent Events (SSE)
# Each client gets its own asyncio.Queue which the POST handler will put new readings into.
listeners: List[asyncio.Queue] = []
excel_lock = threading.Lock()
EXCEL_PATH = Path(__file__).resolve().parent.parent / "data" / "readings.xlsx"


def append_reading_to_excel(row: dict):
    """Append a reading dict to an Excel workbook. Creates file and header if missing.

    row keys expected: timestamp (datetime or str), np_n, np_p, np_k, ph, ec, humidity, temperature
    """
    if not OPENPYXL_AVAILABLE:
        # openpyxl not installed; skip writing to Excel but do not raise.
        logging.warning("openpyxl not available; skipping Excel write for readings")
        return

    EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)
    # Normalize timestamp
    ts = row.get("timestamp")
    try:
        # If it's a string, attempt to split
        if isinstance(ts, str):
            # try to parse ISO-like string
            date_part = ts.split("T")[0] if "T" in ts else ts.split()[0]
            time_part = ts.split("T")[1] if "T" in ts and len(ts.split("T")) > 1 else (ts.split()[1] if len(ts.split()) > 1 else "")
        else:
            date_part = ts.date().isoformat()
            time_part = ts.time().strftime("%H:%M:%S")
    except Exception:
        date_part = str(ts)
        time_part = ""

    headers = ["Time", "Date", "N", "P", "K", "EC", "pH", "Humidity", "Temperature"]
    values = [time_part, date_part, row.get("np_n"), row.get("np_p"), row.get("np_k"), row.get("ec"), row.get("ph"), row.get("humidity"), row.get("temperature")]

    with excel_lock:
        if EXCEL_PATH.exists():
            try:
                wb = load_workbook(EXCEL_PATH)
                ws = wb.active
            except Exception:
                # If loading fails, create a fresh workbook
                wb = Workbook()
                ws = wb.active
                ws.append(headers)
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(headers)

        ws.append(values)
        wb.save(EXCEL_PATH)


async def push_event(data: dict):
    payload = json.dumps(data, default=str)
    # deliver to all listeners; use copy to avoid mutation during iteration
    for q in list(listeners):
        try:
            await q.put(payload)
        except Exception:
            # ignore errors per-listener
            pass



@app.post("/api/readings", status_code=201)
async def post_reading(payload: dict, db: Session = Depends(get_db), _=Depends(require_api_key)):
    # Payload may contain keys: np_n, np_p, np_k, ph, ec, humidity, temperature
    ts_now = datetime.now(timezone.utc)
    r = models.Reading(
        timestamp=ts_now,
        np_n=payload.get("np_n"),
        np_p=payload.get("np_p"),
        np_k=payload.get("np_k"),
        ph=payload.get("ph"),
        ec=payload.get("ec"),
        humidity=payload.get("humidity"),
        temperature=payload.get("temperature"),
        raw=str(payload)
    )
    db.add(r)
    db.commit()
    db.refresh(r)
    # Push event to SSE listeners asynchronously (schedule a background task)
    asyncio.create_task(push_event({
        "id": r.id,
        "timestamp": r.timestamp,
        "np_n": r.np_n,
        "np_p": r.np_p,
        "np_k": r.np_k,
        "ph": r.ph,
        "ec": r.ec,
        "humidity": r.humidity,
        "temperature": r.temperature,
    }))

    # Also append to Excel (run in thread so we don't block the event loop)
    try:
        row = {
            "timestamp": r.timestamp,
            "np_n": r.np_n,
            "np_p": r.np_p,
            "np_k": r.np_k,
            "ph": r.ph,
            "ec": r.ec,
            "humidity": r.humidity,
            "temperature": r.temperature,
        }
        asyncio.create_task(asyncio.to_thread(append_reading_to_excel, row))
    except Exception:
        # best-effort: do not fail the request if excel write fails
        pass
    return {"id": r.id}


@app.get('/api/stream')
def stream_readings(request: Request, api_key: Optional[str] = None, x_api_key: Optional[str] = Header(None)):
    """Server-Sent Events endpoint streaming new readings as they arrive.

    Browsers' EventSource cannot set custom headers, so this endpoint accepts an
    `api_key` query parameter (e.g. `/api/stream?api_key=dev-token`) in addition
    to the normal `x-api-key` header. The server will validate either.
    """
    # Validate provided api key (either query param or header)
    provided = api_key or x_api_key
    if provided is None:
        # If no API key provided, allow EventSource from clients on common
        # private/local networks (loopback and RFC1918 ranges). Browsers
        # cannot set headers on EventSource, and for local/LAN deployments
        # this is a practical convenience to avoid 401s. If you expose this
        # server publicly, consider setting DASHBOARD_API_KEY and not relying
        # on this fallback.
        client_host = getattr(request.client, 'host', None)
        if client_host:
            if (client_host.startswith('127.') or client_host == '::1' or
                    client_host.startswith('192.') or client_host.startswith('10.') or
                    client_host.startswith('172.')):
                logging.info(f"Allowing SSE connection from local client {client_host} without api_key")
                provided = API_KEY  # treat as authorized for this connection
            else:
                raise HTTPException(status_code=401, detail="Missing API key")
        else:
            raise HTTPException(status_code=401, detail="Missing API key")

    if provided != API_KEY:
        raise HTTPException(status_code=403, detail="Invalid API key")

    q: asyncio.Queue = asyncio.Queue()
    listeners.append(q)

    async def event_generator():
        try:
            while True:
                data = await q.get()
                yield f"data: {data}\n\n"
        except asyncio.CancelledError:
            return
        finally:
            # cleanup: remove our queue from listeners
            try:
                listeners.remove(q)
            except ValueError:
                pass

    return StreamingResponse(event_generator(), media_type='text/event-stream')


@app.get("/api/readings/latest")
def latest_reading(db: Session = Depends(get_db), _=Depends(require_api_key)):
    item = db.query(models.Reading).order_by(models.Reading.timestamp.desc()).first()
    if not item:
        raise HTTPException(status_code=404, detail="No readings available")
    return {
        "id": item.id,
        "timestamp": item.timestamp,
        "np_n": item.np_n,
        "np_p": item.np_p,
        "np_k": item.np_k,
        "ph": item.ph,
        "ec": item.ec,
        "humidity": item.humidity,
        "temperature": item.temperature,
    }


@app.get("/api/readings/history")
def history(limit: int = 100, db: Session = Depends(get_db), _=Depends(require_api_key)):
    items = db.query(models.Reading).order_by(models.Reading.timestamp.desc()).limit(limit).all()
    return [
        {
            "id": it.id,
            "timestamp": it.timestamp,
            "np_n": it.np_n,
            "np_p": it.np_p,
            "np_k": it.np_k,
            "ph": it.ph,
            "ec": it.ec,
            "humidity": it.humidity,
            "temperature": it.temperature,
        }
        for it in items
    ]


@app.post("/api/command")
def send_command(cmd: dict, _=Depends(require_api_key)):
    # This endpoint is a placeholder to record commands sent from app/website.
    # Extend to forward commands to a device gateway if present.
    return {"status": "queued", "command": cmd}


@app.post("/api/predict")
def predict_crop(payload: dict, _=Depends(require_api_key)):
    """Predict the most suitable crop from soil values.

    Supported payload keys:
    - np_n or n
    - np_p or p
    - np_k or k
    - ph
    - moisture (fallback to humidity)
    - temperature
    """

    def _to_float(name: str, value: Any) -> float:
        if value is None:
            raise HTTPException(status_code=422, detail=f"Missing required field: {name}")
        try:
            return float(value)
        except Exception:
            raise HTTPException(status_code=422, detail=f"Invalid numeric value for: {name}")

    np_n = _to_float("np_n", payload.get("np_n", payload.get("n")))
    np_p = _to_float("np_p", payload.get("np_p", payload.get("p")))
    np_k = _to_float("np_k", payload.get("np_k", payload.get("k")))
    ph = _to_float("ph", payload.get("ph"))
    moisture = _to_float("moisture", payload.get("moisture", payload.get("humidity")))
    temperature = _to_float("temperature", payload.get("temperature"))

    values = [np_n, np_p, np_k, ph, moisture, temperature]

    try:
        result = _predict_crop(values)
        result["input"] = {
            "np_n": np_n,
            "np_p": np_p,
            "np_k": np_k,
            "ph": ph,
            "moisture": moisture,
            "temperature": temperature,
        }
        return result
    except HTTPException:
        raise
    except RuntimeError as e:
        raise HTTPException(status_code=503, detail=str(e))
    except Exception as e:
        logging.exception("Prediction failed")
        raise HTTPException(status_code=500, detail=f"Prediction error: {e}")


def _predict_history_rows(db: Session, limit: int = 10):
    limit = max(1, min(int(limit), 25))
    items = db.query(models.Reading).order_by(models.Reading.timestamp.desc()).limit(limit).all()
    rows = []

    for item in items:
        try:
            values = [
                float(item.np_n or 0),
                float(item.np_p or 0),
                float(item.np_k or 0),
                float(item.ph or 0),
                float(item.humidity or 0),
                float(item.temperature or 0),
            ]
            result = _predict_crop(values)
            rows.append({
                "id": item.id,
                "timestamp": item.timestamp,
                "np_n": item.np_n,
                "np_p": item.np_p,
                "np_k": item.np_k,
                "ph": item.ph,
                "humidity": item.humidity,
                "temperature": item.temperature,
                "predicted_crop": result["predicted_crop"],
                "confidence": result["confidence"],
                "top_predictions": result["top_predictions"],
            })
        except Exception as e:
            rows.append({
                "id": item.id,
                "timestamp": item.timestamp,
                "np_n": item.np_n,
                "np_p": item.np_p,
                "np_k": item.np_k,
                "ph": item.ph,
                "humidity": item.humidity,
                "temperature": item.temperature,
                "error": str(e),
            })

    return rows


@app.get("/api/predict/history")
def predict_history(limit: int = 10, db: Session = Depends(get_db), _=Depends(require_api_key)):
    """Return predictions for the most recent readings.

    Each row uses the latest reading values as prediction input and returns the
    predicted crop for the page to display in a table.
    """
    return _predict_history_rows(db, limit)


# ============================
# Crop suitability and ideal ranges
# ============================
IDEAL_RANGES = {
    "Chili": {"N": (70, 120), "P": (30, 60), "K": (30, 60), "pH": (5.5, 7.0), "Moisture": (50, 80), "Temperature": (20, 32)},
    "Cucumber": {"N": (60, 120), "P": (20, 50), "K": (30, 80), "pH": (6.0, 7.0), "Moisture": (60, 90), "Temperature": (18, 30)},
    "Groundnut": {"N": (40, 80), "P": (20, 40), "K": (20, 60), "pH": (5.5, 6.5), "Moisture": (40, 70), "Temperature": (20, 30)},
    "Maize": {"N": (80, 140), "P": (30, 80), "K": (30, 80), "pH": (5.5, 7.5), "Moisture": (40, 70), "Temperature": (18, 32)},
    "Paddy": {"N": (80, 150), "P": (40, 80), "K": (40, 100), "pH": (5.0, 6.8), "Moisture": (70, 100), "Temperature": (20, 32)},
    "Spinach": {"N": (40, 100), "P": (20, 60), "K": (30, 70), "pH": (6.0, 7.5), "Moisture": (50, 80), "Temperature": (10, 24)},
}


class SoilEntryCreate(BaseModel):
    label: Optional[str] = None
    n: float
    p: float
    k: float
    ph: float
    moisture: float
    temperature: float


class SuitabilityCheckRequest(BaseModel):
    crop_name: str


class CropTestItem(BaseModel):
    name: str
    values: List[float]


class CropTestRequest(BaseModel):
    crops: List[CropTestItem]


class PlantUpdateRequest(BaseModel):
    plant_name: str


def _suitability_check(desired_crop: str, soil_values: list):
    """Check soil suitability for a crop.
    
    Args:
        desired_crop: Crop name
        soil_values: [N, P, K, pH, Moisture, Temperature]
    
    Returns:
        dict with: suitable (bool), top_prob (float), predicted (str), alternatives (list), procedures (list)
    """
    result = {"suitable": False, "top_prob": 0.0, "predicted": None, "alternatives": [], "procedures": []}
    
    try:
        # Get model prediction
        pred_result = _predict_crop(soil_values)
        predicted_crop = pred_result.get("predicted_crop")
        top_prob = pred_result.get("confidence", 0.0)
        
        result["predicted"] = predicted_crop
        result["top_prob"] = top_prob
        
        # Get alternative crops from top_predictions
        top_preds = pred_result.get("top_predictions", [])
        alternatives = [p["crop"] for p in top_preds if p["crop"].lower() != desired_crop.lower()]
        result["alternatives"] = alternatives[:2]
        
        # Check exact match
        if predicted_crop.lower() == desired_crop.lower() and top_prob >= 0.6:
            result["suitable"] = True
            return result
        
        # If not suitable, provide recommendations based on ideal ranges
        ideal = IDEAL_RANGES.get(desired_crop, {})
        if not ideal:
            result["procedures"] = ["No ideal ranges available for this crop; consider testing soil and improving general fertility."]
            return result
        
        labels = ["N", "P", "K", "pH", "Moisture", "Temperature"]
        procedures = []
        
        for i, label in enumerate(labels):
            if i >= len(soil_values):
                continue
            val = soil_values[i]
            if label in ideal:
                min_val, max_val = ideal[label]
                if val < min_val:
                    procedures.append(f"Increase {label}: currently {val}, should be {min_val}-{max_val}")
                elif val > max_val:
                    procedures.append(f"Decrease {label}: currently {val}, should be {min_val}-{max_val}")
        
        if not procedures:
            procedures = ["Soil matches typical ranges but model doesn't recommend this crop—consider soil biology or other factors."]
        
        result["procedures"] = procedures
        return result
    except Exception as e:
        result["procedures"] = [f"Error checking suitability: {e}"]
        return result


# ============================
# Soil Management Endpoints
# ============================

@app.post("/api/commands/soil")
def add_soil_entry(payload: SoilEntryCreate, db: Session = Depends(get_db), _=Depends(require_api_key)):
    """Add a new soil entry."""
    soil = models.SoilEntry(
        label=payload.label,
        n=payload.n,
        p=payload.p,
        k=payload.k,
        ph=payload.ph,
        moisture=payload.moisture,
        temperature=payload.temperature,
    )
    db.add(soil)
    db.commit()
    db.refresh(soil)
    return {
        "id": soil.id,
        "label": soil.label,
        "values": [soil.n, soil.p, soil.k, soil.ph, soil.moisture, soil.temperature],
        "created_at": soil.created_at,
    }


@app.get("/api/commands/soil")
def list_soil_entries(db: Session = Depends(get_db), _=Depends(require_api_key)):
    """List all saved soil entries."""
    entries = db.query(models.SoilEntry).order_by(models.SoilEntry.created_at.desc()).all()
    return [
        {
            "id": e.id,
            "label": e.label,
            "values": [e.n, e.p, e.k, e.ph, e.moisture, e.temperature],
            "created_at": e.created_at,
        }
        for e in entries
    ]


@app.delete("/api/commands/soil/{soil_id}")
def delete_soil_entry(soil_id: int, db: Session = Depends(get_db), _=Depends(require_api_key)):
    """Delete a soil entry."""
    soil = db.query(models.SoilEntry).filter(models.SoilEntry.id == soil_id).first()
    if not soil:
        raise HTTPException(status_code=404, detail="Soil entry not found")
    db.delete(soil)
    db.commit()
    return {"status": "deleted", "id": soil_id}


@app.post("/api/commands/soil/{soil_id}/suitability")
def check_crop_suitability(soil_id: int, payload: SuitabilityCheckRequest, db: Session = Depends(get_db), _=Depends(require_api_key)):
    """Check if a soil entry is suitable for a crop."""
    soil = db.query(models.SoilEntry).filter(models.SoilEntry.id == soil_id).first()
    if not soil:
        raise HTTPException(status_code=404, detail="Soil entry not found")
    
    values = [soil.n, soil.p, soil.k, soil.ph, soil.moisture, soil.temperature]
    result = _suitability_check(payload.crop_name, values)
    result["soil_id"] = soil_id
    result["soil_label"] = soil.label
    result["crop_name"] = payload.crop_name
    return result


@app.post("/api/commands/soil/test-crops")
def test_all_crops(payload: CropTestRequest, db: Session = Depends(get_db), _=Depends(require_api_key)):
    """Test multiple crops at once.
    
    Expected payload: {"crops": [{"name": "Chili", "values": [N, P, K, pH, Moisture, Temp]}, ...]}
    """
    results = []
    for item in payload.crops:
        try:
            crop_name = item.name
            values = item.values
            if len(values) != 6:
                results.append({
                    "crop": crop_name,
                    "error": "Invalid crop name or values",
                    "suitable": False,
                })
                continue
            
            pred_result = _predict_crop(values)
            predicted = pred_result.get("predicted_crop")
            confidence = pred_result.get("confidence", 0.0)
            
            suitable = (predicted.lower() == crop_name.lower() and confidence >= 0.6)
            results.append({
                "crop": crop_name,
                "values": {"n": values[0], "p": values[1], "k": values[2], "ph": values[3], "moisture": values[4], "temperature": values[5]},
                "predicted": predicted,
                "confidence": confidence,
                "suitable": suitable,
            })
        except Exception as e:
            results.append({
                "crop": item.name,
                "error": str(e),
                "suitable": False,
            })
    
    return {"results": results}


# ============================
# Plant History Endpoints
# ============================

@app.get("/api/commands/plants")
def get_plants(db: Session = Depends(get_db), _=Depends(require_api_key)):
    """Get plant history (previous and next plants)."""
    # For now, we'll maintain a single record (id=1)
    plant_entry = db.query(models.PlantHistory).first()
    if not plant_entry:
        plant_entry = models.PlantHistory(previous_plant=None, next_plant=None)
        db.add(plant_entry)
        db.commit()
        db.refresh(plant_entry)
    
    return {
        "id": plant_entry.id,
        "previous_plant": plant_entry.previous_plant,
        "next_plant": plant_entry.next_plant,
        "updated_at": plant_entry.updated_at,
    }


@app.post("/api/commands/plants/previous")
def set_previous_plant(payload: PlantUpdateRequest, db: Session = Depends(get_db), _=Depends(require_api_key)):
    """Set or update the previous plant."""
    plant_entry = db.query(models.PlantHistory).first()
    if not plant_entry:
        plant_entry = models.PlantHistory(previous_plant=payload.plant_name, next_plant=None)
        db.add(plant_entry)
    else:
        plant_entry.previous_plant = payload.plant_name
    
    db.commit()
    db.refresh(plant_entry)
    return {
        "status": "set",
        "previous_plant": plant_entry.previous_plant,
    }


@app.post("/api/commands/plants/next")
def set_next_plant(payload: PlantUpdateRequest, db: Session = Depends(get_db), _=Depends(require_api_key)):
    """Set or update the next plant."""
    plant_entry = db.query(models.PlantHistory).first()
    if not plant_entry:
        plant_entry = models.PlantHistory(previous_plant=None, next_plant=payload.plant_name)
        db.add(plant_entry)
    else:
        plant_entry.next_plant = payload.plant_name
    
    db.commit()
    db.refresh(plant_entry)
    return {
        "status": "set",
        "next_plant": plant_entry.next_plant,
    }


# ============================
# Status Endpoint
# ============================

@app.get("/api/commands/status")
def get_sensor_status(db: Session = Depends(get_db), _=Depends(require_api_key)):
    """Get latest sensor reading and status."""
    reading = db.query(models.Reading).order_by(models.Reading.timestamp.desc()).first()
    if not reading:
        return {
            "status": "no_data",
            "message": "No sensor readings available",
        }
    
    return {
        "status": "ok",
        "timestamp": reading.timestamp,
        "np_n": reading.np_n,
        "np_p": reading.np_p,
        "np_k": reading.np_k,
        "ph": reading.ph,
        "ec": reading.ec,
        "humidity": reading.humidity,
        "temperature": reading.temperature,
    }


# Serve frontend (PWA) static files from the `dashboard/frontend` directory.
# Mounting static after API routes so `/api/*` remains handled by the app.
frontend_dir = Path(__file__).resolve().parent.parent / "frontend"
if frontend_dir.exists():
    # Mount at root so files like /app.js and /style.css are served as expected.
    app.mount("/", StaticFiles(directory=str(frontend_dir), html=True), name="frontend")
else:
    # If frontend folder missing, root index will return a simple message
    @app.get("/", include_in_schema=False)
    def _root_missing():
        return {"message": "Frontend not found. Place frontend files in dashboard/frontend."}
